"""Export EPA TRI cross-linked data in various formats."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Optional

from rich.console import Console

console = Console()

PROJECT_ROOT = Path(__file__).parent.parent.parent
DEFAULT_EXPORT_DIR = PROJECT_ROOT / "data" / "exports"


def export_data(fmt: str = "all", output_dir: str = "data/exports") -> None:
    """Export data in the specified format(s)."""
    from src.storage.database import get_all_facilities, get_connection

    export_path = PROJECT_ROOT / output_dir
    export_path.mkdir(parents=True, exist_ok=True)

    conn = get_connection()
    facilities = get_all_facilities(limit=100000, conn=conn)
    console.print(f"[dim]Exporting {len(facilities)} facilities...[/dim]")

    # Get releases for summary
    releases = conn.execute("""
        SELECT r.tri_facility_id, r.reporting_year, r.chemical_name, r.cas_number,
               r.carcinogen_flag, r.total_releases_lbs, r.fugitive_air_lbs,
               r.stack_air_lbs, r.water_lbs, r.on_site_release_total,
               r.off_site_release_total
        FROM tri_releases r
        ORDER BY r.tri_facility_id, r.reporting_year
    """).fetchall()
    release_data = [dict(r) for r in releases]

    if fmt in ("csv", "all"):
        _export_csv(facilities, release_data, export_path, conn)
    if fmt in ("json", "all"):
        _export_json(facilities, release_data, export_path, conn)
    if fmt in ("excel", "all"):
        _export_excel(facilities, release_data, export_path, conn)
    if fmt in ("markdown", "all"):
        _export_markdown(facilities, release_data, export_path, conn)

    conn.close()
    console.print(f"[bold green]Exports saved to {export_path}[/bold green]")


def _export_csv(facilities: list[dict], releases: list[dict], export_path: Path, conn=None) -> None:
    """Export to CSV files."""
    # Facilities
    fac_path = export_path / "tri_facilities.csv"
    if facilities:
        fac_cols = [k for k in facilities[0].keys() if k != "id"]
        with open(fac_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fac_cols, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(facilities)
        console.print(f"  [green]CSV: {fac_path} ({len(facilities)} facilities)[/green]")

    # Releases
    rel_path = export_path / "tri_releases.csv"
    if releases:
        rel_cols = list(releases[0].keys())
        with open(rel_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=rel_cols, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(releases)
        console.print(f"  [green]CSV: {rel_path} ({len(releases)} releases)[/green]")

    # Enforcement actions
    if conn:
        enf_rows = conn.execute("""
            SELECT ea.case_number, ea.registry_id, tfl.tri_facility_id,
                   ea.case_name, ea.activity_type, ea.enforcement_type,
                   ea.lead_agency, ea.case_status, ea.settlement_date,
                   ea.penalty_amount, ea.fed_penalty_assessed, ea.enforcement_outcome
            FROM enforcement_actions ea
            LEFT JOIN tri_frs_links tfl ON ea.registry_id = tfl.registry_id
            ORDER BY ea.penalty_amount DESC NULLS LAST
        """).fetchall()
        if enf_rows:
            enf_data = [dict(r) for r in enf_rows]
            enf_path = export_path / "enforcement_actions.csv"
            with open(enf_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=enf_data[0].keys(), extrasaction="ignore")
                writer.writeheader()
                writer.writerows(enf_data)
            console.print(f"  [green]CSV: {enf_path} ({len(enf_data)} enforcement actions)[/green]")

        # Inspections
        insp_rows = conn.execute("""
            SELECT fi.inspection_id, fi.registry_id, tfl.tri_facility_id,
                   fi.program, fi.inspection_type, fi.start_date, fi.end_date,
                   fi.lead_agency, fi.found_violation
            FROM facility_inspections fi
            LEFT JOIN tri_frs_links tfl ON fi.registry_id = tfl.registry_id
            ORDER BY fi.start_date DESC
        """).fetchall()
        if insp_rows:
            insp_data = [dict(r) for r in insp_rows]
            insp_path = export_path / "facility_inspections.csv"
            with open(insp_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=insp_data[0].keys(), extrasaction="ignore")
                writer.writeheader()
                writer.writerows(insp_data)
            console.print(f"  [green]CSV: {insp_path} ({len(insp_data)} inspections)[/green]")

        # Superfund proximity
        sf_rows = conn.execute("""
            SELECT tsp.tri_facility_id, f.facility_name, f.state,
                   ss.site_id, ss.site_name, ss.npl_status,
                   tsp.distance_miles
            FROM tri_superfund_proximity tsp
            JOIN tri_facilities f ON tsp.tri_facility_id = f.tri_facility_id
            JOIN superfund_sites ss ON tsp.site_id = ss.site_id
            ORDER BY tsp.distance_miles
        """).fetchall()
        if sf_rows:
            sf_data = [dict(r) for r in sf_rows]
            sf_path = export_path / "superfund_proximity.csv"
            with open(sf_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=sf_data[0].keys(), extrasaction="ignore")
                writer.writeheader()
                writer.writerows(sf_data)
            console.print(f"  [green]CSV: {sf_path} ({len(sf_data)} proximity records)[/green]")

        # Compliance status
        comp_rows = conn.execute("""
            SELECT cs.registry_id, tfl.tri_facility_id,
                   cs.program, cs.status, cs.status_date,
                   cs.quarters_in_nc as violation_event_count
            FROM compliance_status cs
            LEFT JOIN tri_frs_links tfl ON cs.registry_id = tfl.registry_id
            ORDER BY cs.quarters_in_nc DESC
        """).fetchall()
        if comp_rows:
            comp_data = [dict(r) for r in comp_rows]
            comp_path = export_path / "compliance_status.csv"
            with open(comp_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=comp_data[0].keys(), extrasaction="ignore")
                writer.writeheader()
                writer.writerows(comp_data)
            console.print(f"  [green]CSV: {comp_path} ({len(comp_data)} compliance records)[/green]")

        # RMP facilities
        rmp_fac_rows = conn.execute("""
            SELECT rf.rmp_id, rf.facility_name, rf.street_address, rf.city, rf.state,
                   rf.zip_code, rf.latitude, rf.longitude, rf.frs_registry_id,
                   rf.naics_code, rf.num_processes, rf.num_chemicals,
                   rf.last_submission_date, rf.deregistration_date
            FROM rmp_facilities rf
            ORDER BY rf.state, rf.facility_name
        """).fetchall()
        if rmp_fac_rows:
            rmp_data = [dict(r) for r in rmp_fac_rows]
            rmp_path = export_path / "rmp_facilities.csv"
            with open(rmp_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=rmp_data[0].keys(), extrasaction="ignore")
                writer.writeheader()
                writer.writerows(rmp_data)
            console.print(f"  [green]CSV: {rmp_path} ({len(rmp_data)} RMP facilities)[/green]")

        # RMP accidents
        rmp_acc_rows = conn.execute("""
            SELECT a.rmp_id, rf.facility_name, rf.state,
                   a.accident_date, a.chemical_name, a.release_event,
                   a.deaths_workers, a.deaths_public,
                   a.injuries_workers, a.injuries_public,
                   a.evacuations, a.property_damage_usd
            FROM rmp_accidents a
            JOIN rmp_facilities rf ON a.rmp_id = rf.rmp_id
            ORDER BY (a.deaths_workers + a.deaths_public) DESC, a.property_damage_usd DESC
        """).fetchall()
        if rmp_acc_rows:
            acc_data = [dict(r) for r in rmp_acc_rows]
            acc_path = export_path / "rmp_accidents.csv"
            with open(acc_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=acc_data[0].keys(), extrasaction="ignore")
                writer.writeheader()
                writer.writerows(acc_data)
            console.print(f"  [green]CSV: {acc_path} ({len(acc_data)} RMP accidents)[/green]")

        # TRI-RMP cross-links
        link_rows = conn.execute("""
            SELECT lnk.tri_facility_id, f.facility_name as tri_name, f.state,
                   lnk.rmp_id, rf.facility_name as rmp_name,
                   lnk.link_method, lnk.confidence
            FROM tri_rmp_links lnk
            JOIN tri_facilities f ON lnk.tri_facility_id = f.tri_facility_id
            JOIN rmp_facilities rf ON lnk.rmp_id = rf.rmp_id
            ORDER BY f.state, f.facility_name
        """).fetchall()
        if link_rows:
            link_data = [dict(r) for r in link_rows]
            link_path = export_path / "tri_rmp_links.csv"
            with open(link_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=link_data[0].keys(), extrasaction="ignore")
                writer.writeheader()
                writer.writerows(link_data)
            console.print(f"  [green]CSV: {link_path} ({len(link_data)} TRI-RMP links)[/green]")


def _export_json(facilities: list[dict], releases: list[dict], export_path: Path, conn=None) -> None:
    """Export to JSON files."""
    fac_path = export_path / "tri_facilities.json"
    with open(fac_path, "w") as f:
        json.dump(facilities, f, indent=2, default=str)
    console.print(f"  [green]JSON: {fac_path} ({len(facilities)} facilities)[/green]")

    rel_path = export_path / "tri_releases.json"
    with open(rel_path, "w") as f:
        json.dump(releases, f, indent=2, default=str)
    console.print(f"  [green]JSON: {rel_path} ({len(releases)} releases)[/green]")


def _export_excel(
    facilities: list[dict], releases: list[dict], export_path: Path,
    conn=None,
) -> None:
    """Export to styled Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        console.print("[yellow]openpyxl not installed, skipping Excel export[/yellow]")
        return

    import re
    # Regex to strip illegal XML characters that openpyxl rejects
    _ILLEGAL_XML_RE = re.compile(
        r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
    )

    def _clean_cell(val):
        """Strip illegal XML characters from cell values."""
        if isinstance(val, str):
            return _ILLEGAL_XML_RE.sub('', val)
        return val

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="0984E3", end_color="0984E3", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Facilities sheet
    ws = wb.active
    ws.title = "Facilities"
    if facilities:
        headers = [k for k in facilities[0].keys() if k != "id"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, fac in enumerate(facilities[:50000], 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=_clean_cell(fac.get(header)))

    # Releases sheet
    ws2 = wb.create_sheet("Releases")
    if releases:
        rel_headers = list(releases[0].keys())
        for col_idx, header in enumerate(rel_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, rel in enumerate(releases[:50000], 2):
            for col_idx, header in enumerate(rel_headers, 1):
                ws2.cell(row=row_idx, column=col_idx, value=_clean_cell(rel.get(header)))

    # Enforcement sheet
    if conn:
        enf_rows = conn.execute("""
            SELECT ea.case_number, tfl.tri_facility_id, ea.case_name,
                   ea.enforcement_type, ea.penalty_amount, ea.enforcement_outcome,
                   ea.settlement_date, ea.lead_agency
            FROM enforcement_actions ea
            LEFT JOIN tri_frs_links tfl ON ea.registry_id = tfl.registry_id
            ORDER BY ea.penalty_amount DESC NULLS LAST
            LIMIT 50000
        """).fetchall()
        if enf_rows:
            ws3 = wb.create_sheet("Enforcement")
            enf_data = [dict(r) for r in enf_rows]
            enf_headers = list(enf_data[0].keys())
            for col_idx, header in enumerate(enf_headers, 1):
                cell = ws3.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
            for row_idx, rec in enumerate(enf_data[:50000], 2):
                for col_idx, header in enumerate(enf_headers, 1):
                    ws3.cell(row=row_idx, column=col_idx, value=_clean_cell(rec.get(header)))

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum.cell(row=1, column=1, value="EPA TRI Community Health & Demographics Tracker").font = Font(bold=True, size=14)
    ws_sum.cell(row=2, column=1, value="Built by Nathan Goldberg")
    ws_sum.cell(row=3, column=1, value="nathanmauricegoldberg@gmail.com")
    ws_sum.cell(row=5, column=1, value="Total Facilities").font = Font(bold=True)
    ws_sum.cell(row=5, column=2, value=len(facilities))
    ws_sum.cell(row=6, column=1, value="Total Release Records").font = Font(bold=True)
    ws_sum.cell(row=6, column=2, value=len(releases))

    if conn:
        row = conn.execute("SELECT COUNT(DISTINCT state) as cnt FROM tri_facilities").fetchone()
        ws_sum.cell(row=7, column=1, value="States Covered").font = Font(bold=True)
        ws_sum.cell(row=7, column=2, value=row["cnt"])

        row = conn.execute("SELECT SUM(total_releases_lbs) as t FROM tri_releases WHERE total_releases_lbs IS NOT NULL").fetchone()
        ws_sum.cell(row=8, column=1, value="Total Releases (lbs)").font = Font(bold=True)
        ws_sum.cell(row=8, column=2, value=row['t'] or 0)
        ws_sum.cell(row=8, column=2).number_format = '#,##0'

        row = conn.execute("SELECT COUNT(*) as cnt FROM enforcement_actions").fetchone()
        ws_sum.cell(row=9, column=1, value="Enforcement Actions").font = Font(bold=True)
        ws_sum.cell(row=9, column=2, value=row["cnt"])

        row = conn.execute("SELECT COALESCE(SUM(penalty_amount), 0) as t FROM enforcement_actions").fetchone()
        ws_sum.cell(row=10, column=1, value="Total Penalties ($)").font = Font(bold=True)
        ws_sum.cell(row=10, column=2, value=row['t'] or 0)
        ws_sum.cell(row=10, column=2).number_format = '$#,##0'

        row = conn.execute("SELECT COUNT(*) as cnt FROM facility_inspections").fetchone()
        ws_sum.cell(row=11, column=1, value="Facility Inspections").font = Font(bold=True)
        ws_sum.cell(row=11, column=2, value=row["cnt"])

        row = conn.execute("SELECT COUNT(*) as cnt FROM superfund_sites").fetchone()
        ws_sum.cell(row=12, column=1, value="Superfund NPL Sites").font = Font(bold=True)
        ws_sum.cell(row=12, column=2, value=row["cnt"])

        try:
            row = conn.execute("SELECT COUNT(*) as cnt FROM rmp_facilities").fetchone()
            ws_sum.cell(row=13, column=1, value="RMP Facilities").font = Font(bold=True)
            ws_sum.cell(row=13, column=2, value=row["cnt"])

            row = conn.execute("SELECT COUNT(*) as cnt FROM rmp_accidents").fetchone()
            ws_sum.cell(row=14, column=1, value="RMP Accident Records").font = Font(bold=True)
            ws_sum.cell(row=14, column=2, value=row["cnt"])

            row = conn.execute("SELECT COUNT(*) as cnt FROM tri_rmp_links").fetchone()
            ws_sum.cell(row=15, column=1, value="TRI-RMP Cross-Links").font = Font(bold=True)
            ws_sum.cell(row=15, column=2, value=row["cnt"])
        except Exception:
            pass

        # RMP sheet
        try:
            rmp_rows = conn.execute("""
                SELECT lnk.tri_facility_id, f.facility_name, f.state,
                       rf.rmp_id, rf.facility_name as rmp_name,
                       lnk.link_method, lnk.confidence
                FROM tri_rmp_links lnk
                JOIN tri_facilities f ON lnk.tri_facility_id = f.tri_facility_id
                JOIN rmp_facilities rf ON lnk.rmp_id = rf.rmp_id
                ORDER BY f.state, f.facility_name
                LIMIT 50000
            """).fetchall()
            if rmp_rows:
                ws_rmp = wb.create_sheet("RMP Links")
                rmp_data = [dict(r) for r in rmp_rows]
                rmp_headers = list(rmp_data[0].keys())
                for col_idx, header in enumerate(rmp_headers, 1):
                    cell = ws_rmp.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                for row_idx, rec in enumerate(rmp_data[:50000], 2):
                    for col_idx, header in enumerate(rmp_headers, 1):
                        ws_rmp.cell(row=row_idx, column=col_idx, value=_clean_cell(rec.get(header)))
        except Exception:
            pass

    xlsx_path = export_path / "epa_tri_tracker.xlsx"
    wb.save(xlsx_path)
    console.print(f"  [green]Excel: {xlsx_path}[/green]")


def _export_markdown(
    facilities: list[dict], releases: list[dict], export_path: Path,
    conn=None,
) -> None:
    """Export summary statistics as Markdown."""
    md_path = export_path / "summary_stats.md"
    lines = [
        "# EPA TRI Community Health & Demographics Tracker — Summary Statistics",
        "",
        f"**Total Facilities:** {len(facilities):,}",
        f"**Total Release Records:** {len(releases):,}",
        "",
    ]

    if conn:
        row = conn.execute("SELECT COUNT(DISTINCT state) as cnt FROM tri_facilities").fetchone()
        lines.append(f"**States Covered:** {row['cnt']}")

        row = conn.execute("SELECT COUNT(DISTINCT chemical_name) as cnt FROM tri_releases").fetchone()
        lines.append(f"**Unique Chemicals:** {row['cnt']}")

        row = conn.execute("SELECT SUM(total_releases_lbs) as t FROM tri_releases WHERE total_releases_lbs IS NOT NULL").fetchone()
        total_lbs = row["t"] or 0
        lines.append(f"**Total Releases:** {total_lbs:,.0f} lbs")

        row = conn.execute("SELECT SUM(total_releases_lbs) as t FROM tri_releases WHERE carcinogen_flag = 'YES' AND total_releases_lbs IS NOT NULL").fetchone()
        carc_lbs = row["t"] or 0
        lines.append(f"**Carcinogen Releases:** {carc_lbs:,.0f} lbs")

        # Enforcement stats
        row = conn.execute("SELECT COUNT(*) as cnt FROM enforcement_actions").fetchone()
        lines.append(f"**Enforcement Actions:** {row['cnt']:,}")

        row = conn.execute("SELECT COALESCE(SUM(penalty_amount), 0) as t FROM enforcement_actions").fetchone()
        lines.append(f"**Total Penalties:** ${row['t']:,.0f}")

        row = conn.execute("SELECT COUNT(*) as cnt FROM facility_inspections").fetchone()
        lines.append(f"**Facility Inspections:** {row['cnt']:,}")

        row = conn.execute("SELECT COUNT(*) as cnt FROM superfund_sites").fetchone()
        lines.append(f"**Superfund NPL Sites:** {row['cnt']:,}")

        row = conn.execute("SELECT COUNT(DISTINCT tri_facility_id) as cnt FROM tri_superfund_proximity").fetchone()
        lines.append(f"**Facilities Near Superfund Sites:** {row['cnt']:,}")

        lines.extend(["", "## Top States by Facility Count", ""])
        lines.append("| State | Facilities | Total Releases (lbs) |")
        lines.append("|-------|-----------|---------------------|")
        rows = conn.execute("""
            SELECT f.state, COUNT(DISTINCT f.tri_facility_id) as fac_count,
                   COALESCE(SUM(r.total_releases_lbs), 0) as total_lbs
            FROM tri_facilities f
            LEFT JOIN tri_releases r ON f.tri_facility_id = r.tri_facility_id
            GROUP BY f.state ORDER BY fac_count DESC
        """).fetchall()
        for r in rows:
            lines.append(f"| {r['state']} | {r['fac_count']:,} | {r['total_lbs']:,.0f} |")

        lines.extend(["", "## Top 10 Chemicals by Release Volume", ""])
        lines.append("| Chemical | Total Releases (lbs) | Carcinogen |")
        lines.append("|----------|---------------------|------------|")
        rows = conn.execute("""
            SELECT chemical_name, SUM(total_releases_lbs) as total,
                   CASE WHEN SUM(CASE WHEN carcinogen_flag = 'YES' THEN 1 ELSE 0 END) > 0
                        THEN 'YES' ELSE 'NO' END as carc
            FROM tri_releases WHERE total_releases_lbs IS NOT NULL
            GROUP BY chemical_name ORDER BY total DESC LIMIT 10
        """).fetchall()
        for r in rows:
            lines.append(f"| {r['chemical_name']} | {r['total']:,.0f} | {r['carc']} |")

        lines.extend(["", "## Top 10 Industries by Facility Count", ""])
        lines.append("| Industry Sector | Facilities |")
        lines.append("|----------------|-----------|")
        rows = conn.execute("""
            SELECT industry_sector, COUNT(*) as cnt
            FROM tri_facilities WHERE industry_sector IS NOT NULL
            GROUP BY industry_sector ORDER BY cnt DESC LIMIT 10
        """).fetchall()
        for r in rows:
            lines.append(f"| {r['industry_sector']} | {r['cnt']:,} |")

        # RMP stats
        try:
            row = conn.execute("SELECT COUNT(*) as cnt FROM rmp_facilities").fetchone()
            lines.append(f"**RMP Facilities:** {row['cnt']:,}")

            row = conn.execute("SELECT COUNT(*) as cnt FROM rmp_accidents").fetchone()
            lines.append(f"**RMP Accident Records:** {row['cnt']:,}")

            row = conn.execute("SELECT COUNT(*) as cnt FROM tri_rmp_links").fetchone()
            lines.append(f"**TRI-RMP Cross-Links:** {row['cnt']:,}")

            row = conn.execute("SELECT COALESCE(SUM(deaths_workers + deaths_public), 0) as d, COALESCE(SUM(injuries_workers + injuries_public), 0) as i FROM rmp_accidents").fetchone()
            lines.append(f"**RMP Deaths on Record:** {int(row['d']):,}")
            lines.append(f"**RMP Injuries on Record:** {int(row['i']):,}")
        except Exception:
            pass

        # Enforcement top penalties
        enf_rows = conn.execute("""
            SELECT ea.case_name, ea.enforcement_type, ea.penalty_amount
            FROM enforcement_actions ea
            WHERE ea.penalty_amount > 0
            ORDER BY ea.penalty_amount DESC LIMIT 10
        """).fetchall()
        if enf_rows:
            lines.extend(["", "## Top 10 Enforcement Penalties", ""])
            lines.append("| Case | Type | Penalty ($) |")
            lines.append("|------|------|------------|")
            for r in enf_rows:
                name = (r["case_name"] or "Unknown")[:50]
                lines.append(f"| {name} | {r['enforcement_type'] or 'N/A'} | ${(r['penalty_amount'] or 0):,.0f} |")

    lines.extend([
        "",
        "---",
        "*Built by Nathan Goldberg — nathanmauricegoldberg@gmail.com*",
    ])

    md_path.write_text("\n".join(lines))
    console.print(f"  [green]Markdown: {md_path}[/green]")
